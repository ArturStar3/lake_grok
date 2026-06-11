import random
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from formular.enums import ActionAnimations, Colors
from formular.models import (
    ActionType,
    Country,
    Marker,
    Target,
    TargetAction,
    TargetType,
)

try:
    from openpyxl import load_workbook
except ImportError as exc:
    load_workbook = None
    _OPENPYXL_IMPORT_ERROR = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

SEED_LABEL_PREFIX = "seed:test"

MARKER_ICON_FILES = {
    "Автотоннель": "Автотоннель.svg",
    "Артиллерийский учебный центр": "Артилерийский учебный центр.svg",
    "Аэродром": "Аэродром.svg",
    "База хранения А": "База хранения А.svg",
    "База хранения Б": "База хранения Б.svg",
    "Вирусная опасность": "Вирусная опасность.svg",
    "Гидротехнические сооружения": "Гидротехническое сооружение.svg",
    "Главный учебный центр": "Главный учебный центр.svg",
    "Мост": "Мост.svg",
    "Мотострелковый учебный центр": "Мотострелковый учебный центр.svg",
    "Нефтеперерабатывающий завод": "НПЗ.svg",
    "Оперативное командование": "Оперативное командование.svg",
    "Перевал": "Перевал.svg",
    "Предприятия ОПК": "Предприятия ОПК.svg",
    "ПУ батальона": "ПУ батальона.svg",
    "ПУ бригады": "ПУ бригады.svg",
    "ПУ бригады сокращенного состава": "ПУ кадрированной бригады.svg",
    "ПУ полка": "ПУ полка.svg",
    "Радиационная опасность": "Радиационная опасность.svg",
    "Радиопеленгаторный пункт": "Радиопеленгаторный пункт.svg",
    "Склад хранения": "Склад хранения.svg",
    "Стратегическое командование": "Стратегическое командование.svg",
    "Теле-радиовышка": "Теле-радио вышка.svg",
    "Узловая жд станция": "Узловая жд станция.svg",
    "Центральная база вооружения": "Центральная база вооружения.svg",
    "Центральная база ремонта танков": "Центральная база ремонта танков.svg",
    "Авиабаза": "Авиабаза.svg",
    "Армейский корпус": "Армейский корпус.svg",
    "Дивизия": "Дивизия.svg",
    "Общевойсковая армия": "Общевойсковая армия.svg",
}

COUNTRY_BOUNDS = {
    "CN": (35.0, 48.5, 75.0, 96.0),
    "UZ": (37.5, 45.5, 56.0, 66.5),
    "TJ": (36.5, 41.0, 67.5, 75.5),
    "KZ": (40.5, 54.5, 46.5, 87.0),
    "KG": (39.0, 43.5, 69.0, 80.5),
    "TM": (35.0, 42.5, 52.5, 66.5),
    "AF": (29.5, 38.5, 60.5, 75.0),
    "IR": (25.0, 39.5, 44.0, 63.5),
    "AZ": (38.5, 41.5, 44.5, 51.0),
}

ACTION_TYPES = [
    ("Разведка", ActionAnimations.RADAR),
    ("Патрулирование", ActionAnimations.WAVE),
    ("Огневая поддержка", ActionAnimations.PULSE),
    ("Блокада", ActionAnimations.RINGS),
]


def _parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "да"}


def _parse_xlsx_tables(xlsx_path: Path):
    if load_workbook is None:
        raise CommandError(
            "Для чтения Data.xlsx установите openpyxl: pip install openpyxl"
        ) from _OPENPYXL_IMPORT_ERROR

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    markers = []
    countries = []
    section = None
    headers = None

    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if first in ("formular_marker", "formular_country"):
            section = first
            headers = None
            continue
        if first == "title":
            headers = [cell for cell in row]
            continue
        if not first or not section or not headers:
            continue

        data = dict(zip(headers, row))
        if section == "formular_marker":
            markers.append(data)
        elif section == "formular_country":
            countries.append(data)

    wb.close()

    if not markers:
        raise CommandError("В Data.xlsx не найдена таблица formular_marker")
    if not countries:
        raise CommandError("В Data.xlsx не найдена таблица formular_country")

    return markers, countries


def _project_root():
    return Path(settings.BASE_DIR).parent


def _icons_dir():
    return _project_root() / "Значки"


def _resolve_xlsx_path(explicit_path):
    if explicit_path:
        path = Path(explicit_path)
    else:
        path = _project_root() / "Data.xlsx"
    if not path.exists():
        raise CommandError(f"Файл Data.xlsx не найден: {path}")
    return path


def _random_coord(lat_min, lat_max, lng_min, lng_max, rng):
    return (
        round(rng.uniform(lat_min, lat_max), 6),
        round(rng.uniform(lng_min, lng_max), 6),
    )


class Command(BaseCommand):
    help = (
        "Заполняет Target тестовыми данными на основе "
        "formular_country и formular_marker из Data.xlsx"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--count",
            type=int,
            default=1000,
            help="Количество создаваемых объектов Target (по умолчанию 1000)",
        )
        parser.add_argument(
            "--xlsx",
            type=str,
            default="",
            help="Путь к Data.xlsx (по умолчанию <корень проекта>/Data.xlsx)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help=f"Удалить ранее созданные тестовые Target (label начинается с {SEED_LABEL_PREFIX})",
        )
        parser.add_argument(
            "--seed",
            type=int,
            default=42,
            help="Seed для генератора случайных координат",
        )

    def handle(self, *args, **options):
        count = options["count"]
        if count <= 0:
            raise CommandError("--count должен быть больше 0")

        xlsx_path = _resolve_xlsx_path(options["xlsx"] or None)
        markers_data, countries_data = _parse_xlsx_tables(xlsx_path)
        icons_dir = _icons_dir()
        rng = random.Random(options["seed"])

        with transaction.atomic():
            if options["clear"]:
                deleted, _ = Target.objects.filter(
                    label__startswith=SEED_LABEL_PREFIX
                ).delete()
                self.stdout.write(f"Удалено тестовых Target: {deleted}")

            countries = self._ensure_countries(countries_data)
            marker_records = self._ensure_markers(markers_data, icons_dir)
            action_types = self._ensure_action_types()

            created = self._create_targets(
                count=count,
                countries=countries,
                marker_records=marker_records,
                action_types=action_types,
                rng=rng,
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Готово: создано {created} объектов Target "
                f"({len(countries)} стран, {len(marker_records)} типов маркеров)"
            )
        )

    def _ensure_countries(self, countries_data):
        result = []
        for row in countries_data:
            color = str(row.get("color", Colors.blue.name)).strip()
            if color not in Colors.__members__:
                color = Colors.blue.name

            title = str(row["title"]).strip()
            country, _ = Country.objects.update_or_create(
                title=title,
                defaults={
                    "title_short": str(row["title_short"]).strip(),
                    "iso_code": str(row["iso_code"]).strip(),
                    "color": color,
                },
            )
            result.append(country)
        return result

    def _ensure_markers(self, markers_data, icons_dir):
        result = []
        for row in markers_data:
            title = str(row["title"]).strip()
            target_type, _ = TargetType.objects.get_or_create(title=title)

            marker_defaults = {
                "top": int(row.get("top") or 0),
                "width": int(row.get("width") or 100),
                "height": int(row.get("height") or 50),
                "order": int(row.get("order") or 1),
                "scale": Decimal(str(row.get("scale") or "1.0")),
                "is_flag": _parse_bool(row.get("is_flag", True)),
            }

            marker, created = Marker.objects.get_or_create(
                title=title,
                defaults=marker_defaults,
            )

            if not created:
                Marker.objects.filter(pk=marker.pk).update(**marker_defaults)
                marker.refresh_from_db()

            if not marker.path:
                icon_name = MARKER_ICON_FILES.get(title)
                if not icon_name:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Нет SVG для маркера «{title}», объект будет без иконки"
                        )
                    )
                else:
                    icon_path = icons_dir / icon_name
                    if not icon_path.exists():
                        self.stdout.write(
                            self.style.WARNING(
                                f"Файл не найден: {icon_path}"
                            )
                        )
                    else:
                        with icon_path.open("rb") as svg_file:
                            marker.path.save(icon_name, File(svg_file), save=True)

            result.append(
                {
                    "marker": marker,
                    "target_type": target_type,
                    "title": title,
                    "is_flag": marker.is_flag,
                }
            )

        return result

    def _ensure_action_types(self):
        result = []
        for title, animation in ACTION_TYPES:
            action_type, _ = ActionType.objects.get_or_create(
                title=title,
                defaults={"animation": animation},
            )
            result.append(action_type)
        return result

    def _create_targets(self, count, countries, marker_records, action_types, rng):
        per_country = count // len(countries)
        remainder = count % len(countries)

        targets_to_create = []
        counter = 1

        for country_index, country in enumerate(countries):
            iso = country.iso_code
            bounds = COUNTRY_BOUNDS.get(iso)
            if not bounds:
                raise CommandError(f"Нет координатных границ для страны {iso}")

            country_count = per_country + (1 if country_index < remainder else 0)
            lat_min, lat_max, lng_min, lng_max = bounds

            for _ in range(country_count):
                marker_info = marker_records[(counter - 1) % len(marker_records)]
                lat, lng = _random_coord(lat_min, lat_max, lng_min, lng_max, rng)
                action_radius = (
                    round(rng.uniform(15.0, 180.0), 1)
                    if marker_info["is_flag"] and rng.random() < 0.45
                    else 0.0
                )

                targets_to_create.append(
                    Target(
                        country=country,
                        title=(
                            f"{marker_info['title']} "
                            f"({country.title_short}-{counter:04d})"
                        ),
                        label=f"{SEED_LABEL_PREFIX}:{counter:05d}",
                        marker=marker_info["marker"],
                        type=marker_info["target_type"],
                        action_radius=action_radius or None,
                        lat=lat,
                        lng=lng,
                    )
                )
                counter += 1

        created_targets = Target.objects.bulk_create(
            targets_to_create,
            batch_size=500,
        )

        actions_to_create = []
        for target, source in zip(created_targets, targets_to_create):
            if not source.action_radius:
                continue
            actions_count = 1 if rng.random() < 0.7 else 2
            chosen_types = rng.sample(
                action_types,
                k=min(actions_count, len(action_types)),
            )
            for action_type in chosen_types:
                radius = round(
                    float(source.action_radius) * rng.uniform(0.6, 1.2),
                    1,
                )
                actions_to_create.append(
                    TargetAction(
                        target=target,
                        action_type=action_type,
                        radius=radius,
                    )
                )

        if actions_to_create:
            TargetAction.objects.bulk_create(actions_to_create, batch_size=500)

        return len(created_targets)